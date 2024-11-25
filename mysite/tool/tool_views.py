from django.shortcuts import get_object_or_404, render, redirect
from django.http import HttpResponse, HttpResponseRedirect, Http404, JsonResponse, FileResponse, StreamingHttpResponse
from django.conf import settings
from rest_framework.decorators import api_view, parser_classes
from rest_framework.parsers import MultiPartParser
from datetime import datetime, timedelta
import pytz #時間區域
from dateutil.parser import parse #時間字串改回時間 用於時間比較
import uuid #亂數碼
import json, os, time
import pickle
import base64
import inspect #取得對象訊息/代碼可讀
import openpyxl
import tempfile
#清除緩存
from django.views.decorators.cache import cache_control
#跨站偽造
from django.views.decorators.csrf import csrf_exempt

from django.db import connection
from user.token import LoginRequiredMiddleware #request.uesr_id
from sharepoint import sharepoint_views as sharepoint_views
from log import log_views as log_views
from user import user_views as user_views

class DateTimeEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, datetime):
            return obj.isoformat()
        return super().default(obj)
    
def timenow():
    current_time = datetime.now()
    timezone = pytz.timezone('Asia/Taipei')
    last_update_time = current_time.astimezone(timezone)
    return last_update_time    

def with_db_connection(func):
    def wrapper(*args, **kwargs):
        with connection.cursor() as cursor:
            return func(cursor, *args, **kwargs)
    return wrapper

@api_view(["post"])
@csrf_exempt
@with_db_connection 
def select_long_name(cursor, request):
    shortname = request.data.get('shortname')
    if shortname:
        shortname_set = set(shortname)  #set送進來的數值  要是["ABC"] 才會變成 {"ABC"}   /    如果送ABC就會變成{"A","B","C"}
        shortname_condition = f"AND dtl.short_name IN ({', '.join(['%s'] * len(shortname_set))})"
    else:
        shortname_set = set("")
        shortname_condition = ""
    query = f'''
    SELECT DISTINCT long_name
    FROM device_tool_list AS dtl
    WHERE 1=1 {shortname_condition}
    '''
    cursor.execute(query, (tuple(shortname_set)))
    rows = cursor.fetchall()
    long_name = {'long_name': [row[0] for row in rows]}
    return JsonResponse(long_name)
    
@api_view(["post"])
@csrf_exempt
@parser_classes([MultiPartParser])
@with_db_connection
def download_version(cursor, request):
    user = request.user_id
    version = json.loads(request.data.get('version'))
    file_name =f'{version}.zip'
    try:
        temp_dir = sharepoint_views.sharepoint_download_file_zip(user, file_name)
        temp_file_path = os.path.join(temp_dir, file_name)
        #download_url = f'/pulsar/download_file/{file_name}/'
        print(temp_file_path, temp_dir)
        file_size = os.path.getsize(temp_file_path)
        print(file_size)
        # with open(temp_file_path, 'rb') as file:
        #     response = FileResponse(file)
        # FileResponse(open(temp_file_path, 'rb'), content_type='application/octet-stream')
        # response = FileResponse(open(temp_file_path, 'rb'), content_type='application/octet-stream')
        # response['Content-Disposition'] = f'attachment; filename="{os.path.basename(temp_file_path)}"'
        # return JsonResponse({'finaldata': 'successful download'})
        # response = StreamingHttpResponse(file_iterator(temp_file_path), content_type='application/octet-stream')
        # response['Content-Disposition'] = f'attachment; filename="{os.path.basename(temp_file_path)}"'
        
        # # 在文件流式傳送完成後，執行後續操作
        # response.streaming_content = cleanup_temp_files(temp_file_path, temp_dir)

        # def cleanup_temp_files(temp_file_path, temp_dir):
        #     os.remove(temp_file_path)
        #     os.rmdir(temp_dir)
        # cleanup_queue_new.append([temp_file_path, temp_dir])
        return FileResponse(open(temp_file_path, 'rb'), content_type='application/octet-stream')
    except Exception as e:
        print(e)
        log_views.log_error(__file__, inspect.currentframe().f_code.co_name, e)
        return JsonResponse({'error': '下載失敗 請聯絡管理員'})

def download_file_exe(file_name):
    import requests
    from requests.auth import HTTPBasicAuth
    from O365.utils import BaseTokenBackend
    import os, sys
    from pathlib import Path
    from O365 import Account
    SERVER_HOST_NAME = "env-lab.eba-jpevj6xq.ap-southeast-1.elasticbeanstalk.com"
    class JustToken(BaseTokenBackend):

        def __init__(self):
            super().__init__()

        def load_token(self):
            if self.token:
                return self.token
            return None

        def save_token(self):
            return self.token

        def check_token(self):
            return self.token is not None


    class sharepoint:
        def __init__(self):
            self.__set_token_database_config()
            self.token_backend = self.get_token_backend()
            self.token = self.token_backend.load_token()
            self.account = self.get_account
        def __set_token_database_config(self):
            self.server = SERVER_HOST_NAME
            self.token_url = f'http://{self.server}/api/mstoken/6'

        @property
        def get_credential(self):
            try:
                token_url = f'http://{self.server}/api/mstoken/6'
                token_data = requests.get(token_url)
                credentials = None
                if token_data.status_code == 200:
                    token_data = token_data.json()
                    credentials = (token_data['credentials']['appid'],token_data['credentials']['secret'])
                if credentials:
                    return credentials
            except:
                return None

        def __get_token_from_database(self):
            try:
                r = requests.get(self.token_url,
                                auth=HTTPBasicAuth('admin', 'admin'))
                r = r.json()
                return r
            except Exception as e:
                print(e)
            return None
        
        def get_token_backend(self):
            database_token = self.__get_token_from_database()
            database_token_backend = JustToken()
            database_token_backend.token = database_token['token']
            database_token_backend.load_token()
            self.token = database_token_backend.token
            return database_token_backend
        @property
        def curdir(self):
            if getattr(sys, 'frozen', False):
                return Path(sys.executable).parent
            elif __file__:
                return Path(__file__).parent
                
        @property
        def get_account(self):
            credential = self.get_credential
            if self.token_backend.check_token():
                print("完成驗證account")
                return Account(credentials=credential,
                            token_backend=self.token_backend)
            else:
                if credential:
                    account = Account(credential)
                    if account.authenticate(scopes=[
                            'basic', 'users', 'address_book', 'message_all',
                            'onedrive_all', 'sharepoint'
                    ]):
                        print('Authenticated')
                        return account
                else:
                    print('Authenticate Fail !')
                    return None

        @property
        def sharepoint(self):
            return self.account.sharepoint() if self.account else None

        @property
        def commsite(self):
            ret = self.sharepoint.search_site('comm tech')
            if len(ret) > 0: return ret[0]
            else: None
        @property
        def get_validation_lib(self):
            if self.account:
                validation_lib = [
                    lib for lib in self.commsite.list_document_libraries()
                    if lib.name == 'Validation and Quality Program'
                ][0]
                return validation_lib
            return None
        
        def get_folder_under_validation_lib(self, path):
            return self.get_validation_lib.get_item_by_path(path)

        def upload_to_path(self, path: str, file_path: str):
                try:
                    print(self.get_validation_lib)
                    drive_path = self.get_validation_lib.get_item_by_path(path)
                    filename = os.path.basename(file_path)
                    with open(file_path, "rb") as file:
                        drive_path.upload_file(file_path, filename)
                except Exception as e:        
                    print("Error uploading file:", e)

        def download_from_sharepoint(self, path: str, name: str, to_path=None):
            updater_binery = [
                i for i in self.get_validation_lib.get_item_by_path(path).get_items() if i.name == name
            ]
            if updater_binery:
                path = self.curdir.absolute().as_posix()
                if to_path:
                    path = to_path
                updater_binery[0].download(path)

    sharepoint_download = sharepoint()
    folder_path = f'/IUR/test_folder_Bill'
    sharepoint_download.download_from_sharepoint(folder_path, file_name)



# cleanup_queue_new = []
# cleanup_queue = []    
# import threading, time    
# def cleanup_temp_files():
#     global cleanup_queue_new
#     global cleanup_queue
#     while True:
#         if cleanup_queue:
#             for queue in cleanup_queue:
#                 temp_file_path, temp_dir = queue
#                 os.remove(temp_file_path)
#                 os.rmdir(temp_dir)
#         cleanup_queue = cleanup_queue_new
#         cleanup_queue_new = []
#         time.sleep(60)  

# cleanup_thread = threading.Thread(target=cleanup_temp_files)
# cleanup_thread.start()

def excel_make(data):
    temp_dir = tempfile.mkdtemp()
    file_name = "temp_excel.xlsx"
    file_path = os.path.join(temp_dir, file_name)
    print(temp_dir, file_path)
    wb = openpyxl.Workbook()
    sheet = wb[wb.sheetnames[0]]
    headers = ["platform", "phase", "target", "group", "cycle", "sku", "sn", "borrower", "status", "position", "remark", "update_time"]
    for index, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=index).value = header
    for index, data_item in enumerate(data, start=2):
        sheet.cell(row=index, column=1).value = data_item["platform"]
        sheet.cell(row=index, column=2).value = data_item["phase"]
        sheet.cell(row=index, column=3).value = data_item["target"]
        sheet.cell(row=index, column=4).value = data_item["group"]
        sheet.cell(row=index, column=5).value = data_item["cycle"]
        sheet.cell(row=index, column=6).value = data_item["sku"]
        sheet.cell(row=index, column=7).value = data_item["sn"]
        sheet.cell(row=index, column=8).value = data_item["borrower"]
        sheet.cell(row=index, column=9).value = data_item["status"]
        sheet.cell(row=index, column=10).value = data_item["position"]
        sheet.cell(row=index, column=11).value = data_item["remark"]
        sheet.cell(row=index, column=12).value = data_item["update_time"]
    wb.save(file_path)
    return file_path

def time_transmit(request):
    current_time = datetime.now()
    timezone = pytz.timezone('Asia/Taipei')
    last_update_time = current_time.astimezone(timezone)
    print(last_update_time)
    return JsonResponse({'time': last_update_time})